# import csv

# from suffix_trees import STree

# from constants import *
import json

import spacy

from collections import defaultdict

from openpyxl import load_workbook
from tqdm import tqdm

from merge_exercises import write_examples_to_excel

# nlp = spacy.load("ro_core_news_lg", disable=["parser", "ner", "lemmatizer", "textcat"])
nlp = spacy.load("ro_core_news_lg", disable=["ner", "lemmatizer", "textcat"])


def fix_diacritics(text):
    return text.replace('ş', 'ș').replace('Ş', 'Ș').replace('ţ', 'ț').replace('Ţ', 'Ț')


def read_labels(file_name, include_colored_labels, skip_ambiguous_labels):
    DEFAULT_COLOR = '00000000'
    wb_labels = load_workbook(file_name, read_only=True)
    ws_labels = wb_labels.worksheets[0]
    labeled_verbs = {}
    for i, row in enumerate(ws_labels.iter_rows(min_row=2)):
        # if i == 0:
        #     continue
        if len(row) < 2:
            continue
        verb, label = row[0], row[1]
        if verb.fill and verb.fill.bgColor.rgb != DEFAULT_COLOR and not include_colored_labels:
            continue
        if not label or not label.value:
            continue
        if '/' in label.value and skip_ambiguous_labels:
            continue
        fixed_label = label.value.strip()
        if fixed_label in ['undertand', 'undesrstand', 'undestand']:
            fixed_label = 'understand'
        labeled_verbs[verb.value] = fixed_label
    wb_labels.close()
    return labeled_verbs


def label_example(exercise, labeled_verbs):
    local_tally = defaultdict(lambda: 0)
    selected_verbs = []
    for i, token in enumerate(exercise):
        if str(token) == '-ți':
            selected_verb = (str(exercise[i - 1]) + str(token)).lower()
        else:
            selected_verb = str(token).lower()
        label = labeled_verbs.get(selected_verb)
        if label:
            local_tally[label] += 1
            selected_verbs.append(selected_verb)
    if not local_tally:
        all_verbs = sum([1 for token in exercise if token.pos_ in ('VERB', 'AUX')])
        if all_verbs and len(exercise) > 30:
            return '', local_tally, [], 'NO_LABEL'
        else:
            return '', local_tally, [], 'NO_VERB'
    max_cat = max(local_tally, key=lambda cat: local_tally[cat])
    return max_cat, local_tally, selected_verbs, None


def label_examples(include_colored_labels=False, skip_ambiguous_labels=True, skip_ambiguous_exercises=True, sentence_split=False, remove_ambiguities=False):
    labeled_verbs = {}
    all_labels = defaultdict(lambda: [])
    label_files = ['Verbe-cu-etichetă-p1.xlsx', 'Verbe-cu-etichetă-p2.xlsx', 'Verbe-cu-etichetă-p3.xlsx']
    all_verbs = []
    for label_file in label_files:
        labeled_verbs_batch = read_labels(label_file, include_colored_labels, skip_ambiguous_labels)
        for verb, label in labeled_verbs_batch.items():
            new_verb = fix_diacritics(verb)
            # do not add the same verb twice in the same category
            if new_verb not in all_labels[label]:
                all_labels[label].append(new_verb)
                all_verbs.append(new_verb)
        # NOTE: we only take the last label for ambiguous verbs
        labeled_verbs.update(labeled_verbs_batch)

    labels = defaultdict(lambda: [])
    for verb in labeled_verbs:
        labels[labeled_verbs[verb]].append(verb)
    # print(f"All verbs: {len(all_verbs)}")
    # print(f"Unique verbs: {len(set(all_verbs))}")
    print(f"Duplicate verbs: {set(verb for verb in all_verbs if all_verbs.count(verb) > 1)}")
    suffix = ''
    if include_colored_labels:
        suffix += '_colored'
    kw_file = f"keywords{suffix}.json"
    # print(labeled_verbs)
    # return
    for level in labels:
        # sort labels to easier spot duplicates
        labels[level] = sorted(set(labels[level]))
    with open(kw_file, "w", encoding='utf-8') as f:
        json.dump(labels, f, ensure_ascii=False) # , indent=2)
    # return
    if not skip_ambiguous_exercises:
        if remove_ambiguities:
            suffix += '_break_ties'
        else:
            suffix += '_with_ties'
    if sentence_split:
        suffix += '_split'
    ambiguous_exercises_count = 0
    broken_ties = 0
    labeled_examples = []
    wb_examples = load_workbook('all_exercises_merged.xlsx')
    # wb_examples = load_workbook(f'all_exercises_nonlabeled{suffix}.xlsx')
    ws_examples = wb_examples.worksheets[0]
    skipped_exercises = 0
    skipped_no_verbs = 0
    total_exercises = 0
    non_labeled_examples = []
    for i, row in tqdm(list(enumerate(ws_examples.values))):
        if i == 0:
            continue
        # if i > 50:
        #     break
        #     exit(0)
        id_, publisher, klass, chapter, page, bloom_label, raw_exercise = row
        # tokenize example
        if not raw_exercise:
            # print(f"Empty exercise: {i}")
            continue
        doc = nlp(raw_exercise)
        if sentence_split:
            if remove_ambiguities:
                # in each sentence, use dependency parsing
                #   for each root
                #     (root.children if child.pos_ in ('VERB', 'AUX')) to identify other main verbs
                #     extract each such sub-sentence with token.subtree and create a new example
                #     for sub-sentences that still have multiple labels, take the label from the first verb
                exercises = []
                for exercise in doc.sents:
                    roots = [token for token in exercise if token.dep_ == 'ROOT']
                    for root in roots:
                        root_subtree = list(root.subtree)
                        root_left = root_subtree[0].i
                        root_right = root_subtree[-1].i
                        children = [child for child in root.children if child.pos_ in ('VERB', 'AUX')]
                        # if children:
                        # spans = [list(child.subtree) for child in children]
                        child_tokens = [token for child in children for token in child.subtree]
                        # bounds = [(span[0].i, span[-1].i) for span in spans]
                        root_subtree = [token for token in root_subtree if token not in child_tokens]
                        for child in children:
                            exercises.append([list(child.subtree), exercise.text])
                        exercises.append([root_subtree, exercise.text])
            else:
                exercises = doc.sents
        else:
            exercises = [doc]
        for exercise_init in exercises:
            if remove_ambiguities:
                exercise, exercise_raw = exercise_init
            else:
                exercise = exercise_init
            bloom_label = ''
            if sentence_split:
                if remove_ambiguities:
                    exercise_text = ' '.join([token.text for token in exercise])
                else:
                    exercise_text = exercise.text
                    exercise_raw = exercise_text
            else:
                exercise_text = exercise.text
                exercise_raw = exercise_text
            total_exercises += 1
            max_cat, local_tally, selected_verbs, label_error = label_example(exercise, labeled_verbs)
            if not max_cat:
                if label_error == 'NO_LABEL':
                    skipped_exercises += 1
                    non_labeled_examples.append([publisher, klass, chapter, page, bloom_label, exercise_text])
                else: # if label_error == 'NO_VERB'
                    skipped_no_verbs += 1
                continue
            max_cats = [cat for cat in local_tally if local_tally[cat] == local_tally[max_cat]]
            if len(max_cats) > 1:
                if remove_ambiguities:
                    bloom_label = labeled_verbs.get(str(selected_verbs[0]).lower())
                    broken_ties += 1
                else:
                    ambiguous_exercises_count += 1
                    if skip_ambiguous_exercises:
                        continue
                    bloom_label = '/'.join(sorted(max_cats))
            else:
                bloom_label = max_cat
            # print(f"label {exercise_text}\n\twith {bloom_label}")
            if sentence_split:
                labeled_examples.append([publisher, klass, chapter, page, bloom_label, exercise_text, exercise_raw])
            else:
                labeled_examples.append([publisher, klass, chapter, page, bloom_label, exercise_text])
    print(f"Exercises skipped (no relevant verbs/no verbs or too short): {skipped_exercises} ({skipped_no_verbs}) / {total_exercises}")
    print(f"Ambiguous examples: {ambiguous_exercises_count}, broken ties: {broken_ties}")
    # return
    out_file = f"all_exercises_labeled{suffix}.xlsx"
    write_examples_to_excel(out_file, labeled_examples, sentence_split)
    out_file = f"all_exercises_nonlabeled{suffix}.xlsx"
    write_examples_to_excel(out_file, non_labeled_examples)
    wb_examples.close()


if __name__ == "__main__":
    label_examples(include_colored_labels=False, skip_ambiguous_exercises=True)
    label_examples(include_colored_labels=True, skip_ambiguous_exercises=True)
    label_examples(include_colored_labels=False, skip_ambiguous_exercises=False, sentence_split=True, remove_ambiguities=True)
    label_examples(include_colored_labels=True, skip_ambiguous_exercises=False, sentence_split=True, remove_ambiguities=True)
    label_examples(include_colored_labels=False, skip_ambiguous_exercises=False, sentence_split=True, remove_ambiguities=False)
    label_examples(include_colored_labels=True, skip_ambiguous_exercises=False, sentence_split=True, remove_ambiguities=False)
    label_examples(include_colored_labels=False, skip_ambiguous_exercises=False, remove_ambiguities=False)
    label_examples(include_colored_labels=True, skip_ambiguous_exercises=False, remove_ambiguities=False)
